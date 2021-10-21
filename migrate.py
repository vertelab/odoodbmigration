import sys
import odoorpc
import openpyxl
from pathlib import Path
from mapping import *


IMPORT = '__import__'


def main(model, file):
    print("Calling script with")
    print(f"model = {model}")
    print(f"file  = {file}")
    if input('Continue? ').lower() == 'y':
        migrate_model_from_file(model, file)


def migrate_model_from_file(model, file):
    xlsx_file = Path(file)
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    fkeys = [x.value for x in ws[1]]
    m = map.get(model)
    count = 0
    for row in ws.iter_rows(min_row=2):
        vals = {}
        for key in m:
            if m[key] in fkeys:
                i = fkeys.index(m[key])
                vals.update({key: row[i].value})
        create_record_and_xmlid(model, vals, vals.pop('ext_id'))


def create_record_and_xmlid(model, vals, ext_id):
    """Creates record if it doesn't exist and creates an xml_id"""
    xml_id = generate_xml_id(model, ext_id)
    res_id = get_res_id_from_xml_id(xml_id)
    if res_id:
        print(f"xml_id already exist ({res_id} {xml_id})")
    else:
        try:
            res_id = target.env[model].create(vals)
        except Exception as e:
            print('CREATE_RECORD: FAIL! Read the log...', e, vals)
        else:
            print(f"CREATE_RECORD: SUCCESS! Creating xml_id...")

            try:
                create_xml_id(model, res_id, xml_id)
            except Exception as e:
                print("CREATE_XML_ID: FAIL! Should not happen..."
                      "Did you call this method manually? ", vals)
            else:
                print(f"CREATE_XML_ID: SUCCESS! {xml_id}")
        finally:
            return res_id


def generate_xml_id(model, ext_id):
    """Returns xml_id generated from model and ext_id"""
    return f"{IMPORT}.{model.replace('.', '_')}_{ext_id}"


def get_res_id_from_xml_id(xml_id):
    """Returns res id from xml_id if found."""
    return target.env['ir.model.data'].xmlid_to_res_id(xml_id)


def create_xml_id(model, res_id, xml_id):
    """Creates an xml_id for a record."""
    vals = {'model': model,
            'module': xml_id.split('.')[0],
            'name': xml_id.split('.')[1],
            'res_id': res_id,
            }
    target.env['ir.model.data'].create(vals)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(
            f"\nUsage: python3 {sys.argv[0]} res.partner /path/of/file.xslx\n")
        raise SyntaxError("Wrong number of argument.")

    try:
        target = odoorpc.ODOO().load('target')
    except Exception as e:
        print(e)

    print(target.env)
    main(sys.argv[1], sys.argv[2])
